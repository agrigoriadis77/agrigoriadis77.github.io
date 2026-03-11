#!/usr/bin/env python3
"""
xls_to_schedule_data.py
-----------------------
Converts a .xlsx / .xls school schedule file into js/schedule_data.js,
replacing the old workflow of uploading the file in the browser and
copy-pasting the `defaultData` value via the debugger.

Usage:
    python3 xls_to_schedule_data.py path/to/schedule.xlsx

The script will overwrite  ../js/schedule_data.js  relative to its own
location (i.e.  <repo>/js/schedule_data.js).

Requirements:
    pip install openpyxl xlrd
  (openpyxl handles .xlsx; xlrd handles legacy .xls)
"""

import sys
import os
import json

def read_xls(path: str) -> list[list]:
    """Read the first sheet of an xlsx/xls file and return rows as lists."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            sys.exit(
                "❌  openpyxl is not installed.\n"
                "    Set up the venv once:\n"
                "        python3 -m venv ~/scheduleenv\n"
                "        ~/scheduleenv/bin/pip install openpyxl xlrd\n"
                "    Then run the script with:\n"
                "        ~/scheduleenv/bin/python3 script/xls_to_schedule_data.py ..."
            )
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows():
            rows.append([str(cell.value).strip() if cell.value is not None else None for cell in row])
        return rows

    elif ext == ".xls":
        try:
            import xlrd
        except ImportError:
            sys.exit(
                "❌  xlrd is not installed.\n"
                "    Set up the venv once:\n"
                "        python3 -m venv ~/scheduleenv\n"
                "        ~/scheduleenv/bin/pip install openpyxl xlrd\n"
                "    Then run the script with:\n"
                "        ~/scheduleenv/bin/python3 script/xls_to_schedule_data.py ..."
            )
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                # xlrd types: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
                if cell.ctype == 0:
                    row.append(None)
                elif cell.ctype == 2:
                    # numbers: convert to int if whole, else str
                    v = cell.value
                    row.append(str(int(v)) if v == int(v) else str(v))
                else:
                    row.append(str(cell.value).strip() or None)
            rows.append(row)
        return rows
    else:
        sys.exit(f"❌  Unsupported file type: {ext}  (expected .xlsx or .xls)")


def pad_rows(rows: list[list]) -> list[list]:
    """Pad all rows to the same length with None."""
    max_len = max((len(r) for r in rows), default=0)
    for row in rows:
        while len(row) < max_len:
            row.append(None)
    return rows


def rows_to_js(rows: list[list]) -> str:
    """Serialise rows as a JS const defaultData = [...] assignment."""
    json_str = json.dumps(rows, ensure_ascii=False, indent=4)
    return f"const defaultData = {json_str};\n"


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 xls_to_schedule_data.py path/to/schedule.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.isfile(input_path):
        sys.exit(f"❌  File not found: {input_path}")

    # Output goes to <repo>/js/schedule_data.js regardless of cwd
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "..", "js", "schedule_data.js")
    output_path = os.path.normpath(output_path)

    print(f"📖  Reading: {input_path}")
    rows = read_xls(input_path)
    rows = pad_rows(rows)
    js_content = rows_to_js(rows)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"✅  schedule_data.js written to: {output_path}")
    print(f"    Rows: {len(rows)},  Cols: {len(rows[0]) if rows else 0}")
    print()
    print("Next steps:")
    print("  1. Commit & push the updated js/schedule_data.js")
    print("  2. Done — no browser debugger needed!")


if __name__ == "__main__":
    main()

